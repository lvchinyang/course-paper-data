import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =====================================
# 1. 文件路径设置（Mac 桌面）
# =====================================
desktop = os.path.expanduser("~/Desktop")

input_file = os.path.join(desktop, "菌草技术 25国数据库.xlsx")  # 注意空格和文件名一致
output_file = os.path.join(desktop, "菌草技术25国_描述统计与交叉表.xlsx")

sheet_name = "country_case_database"

if not os.path.exists(input_file):
    raise FileNotFoundError(f"没有找到文件: {input_file}. 请确认文件名和路径是否正确。")

# =====================================
# 2. 读取 Excel
# =====================================
df = pd.read_excel(input_file, sheet_name=sheet_name)
df.columns = df.columns.str.strip()
df = df.fillna("NA")

# =====================================
# 3. 工具函数
# =====================================
def split_multi_value(series):
    values = []
    for item in series:
        if pd.isna(item) or str(item).strip() == "" or str(item).strip().upper() == "NA":
            continue
        item = str(item).replace("；", ";").replace("、", ";").replace("，", ";").replace(",", ";")
        parts = [x.strip() for x in item.split(";") if x.strip() and x.strip().upper() != "NA"]
        values.extend(parts)
    return values

def frequency_table(df, column, table_name):
    total = len(df)
    result = df[column].replace("", "NA").value_counts(dropna=False).reset_index()
    result.columns = [column, "频数"]
    result["占比"] = (result["频数"] / total).round(4)
    result.insert(0, "统计表", table_name)
    return result

def multi_frequency_table(df, column, table_name):
    values = split_multi_value(df[column])
    if len(values) == 0:
        return pd.DataFrame(columns=["统计表", column, "频数", "占比"])
    temp = pd.Series(values).value_counts().reset_index()
    temp.columns = [column, "频数"]
    total_cases = len(df)
    temp["占比"] = (temp["频数"] / total_cases).round(4)
    temp.insert(0, "统计表", table_name)
    return temp

def explode_multi_column(df, column):
    temp = df.copy()
    temp[column] = temp[column].astype(str)
    temp[column] = temp[column].replace("；", ";").replace("、", ";").replace("，", ";").replace(",", ";")
    temp[column] = temp[column].apply(
        lambda x: [i.strip() for i in x.split(";") if i.strip() and i.strip().upper() != "NA"]
    )
    temp = temp.explode(column)
    temp = temp[temp[column].notna()]
    temp = temp[temp[column] != ""]
    temp = temp.reset_index(drop=True)  # 修正重复索引问题
    return temp

def cross_table_one_multi(df, row_col, multi_col):
    temp = explode_multi_column(df, multi_col).copy()
    if temp.empty:
        return pd.DataFrame()
    temp = temp[[row_col, multi_col]].copy()
    ct = pd.crosstab(temp[row_col], temp[multi_col])
    ct["合计"] = ct.sum(axis=1)
    total_row = pd.DataFrame(ct.sum(axis=0)).T
    total_row.index = ["合计"]
    ct = pd.concat([ct, total_row])
    return ct

def cross_table_two_multi(df, multi_col_1, multi_col_2):
    rows = []
    for _, row in df.iterrows():
        list_1 = split_multi_value(pd.Series([row[multi_col_1]]))
        list_2 = split_multi_value(pd.Series([row[multi_col_2]]))
        for a in list_1:
            for b in list_2:
                rows.append([a, b])
    if len(rows) == 0:
        return pd.DataFrame()
    temp = pd.DataFrame(rows, columns=[multi_col_1, multi_col_2])
    ct = pd.crosstab(temp[multi_col_1], temp[multi_col_2])
    ct["合计"] = ct.sum(axis=1)
    total_row = pd.DataFrame(ct.sum(axis=0)).T
    total_row.index = ["合计"]
    ct = pd.concat([ct, total_row])
    return ct

def make_key_case_table(df):
    temp = df.copy()
    def is_number(x):
        try:
            if str(x).strip().upper() == "NA":
                return 0
            float(x)
            return 1
        except:
            return 0
    def mechanism_count(x):
        return len(set(split_multi_value(pd.Series([x]))))
    def evidence_score(x):
        try:
            return int(float(x))
        except:
            return 0
    temp["training_count_是否有数字"] = temp["training_count"].apply(is_number)
    temp["household_count_是否有数字"] = temp["household_count"].apply(is_number)
    temp["机制数量"] = temp["mechanism_type"].apply(mechanism_count)
    temp["案例筛选得分"] = 0
    temp["案例筛选得分"] += temp["evidence_strength"].apply(evidence_score) * 3
    temp["案例筛选得分"] += temp["case_level"].apply(lambda x: 2 if x=="full_country_case" else 1)
    temp["案例筛选得分"] += temp["training_count_是否有数字"]
    temp["案例筛选得分"] += temp["household_count_是否有数字"]
    temp["案例筛选得分"] += temp["机制数量"]
    temp = temp.sort_values(by="案例筛选得分", ascending=False)
    result = temp[[
        "case_id","country_cn","country_en","region","case_level","training_count",
        "household_count","mechanism_type","evidence_strength","案例筛选得分","notes"
    ]].copy()
    return result

# =====================================
# 4. 生成频数表
# =====================================
region_freq = frequency_table(df, "region", "地区分布")
case_level_freq = frequency_table(df, "case_level", "案例层级分布")
evidence_freq = frequency_table(df, "evidence_strength", "证据强度分布")
project_type_freq = multi_frequency_table(df, "project_type", "项目类型频数")
beneficiaries_freq = multi_frequency_table(df, "beneficiaries", "受益对象频数")
outcome_type_freq = multi_frequency_table(df, "outcome_type", "成效类型频数")
mechanism_type_freq = multi_frequency_table(df, "mechanism_type", "机制类型频数")
small_beautiful_freq = multi_frequency_table(df, "small_beautiful_feature", "小而美特征频数")

# =====================================
# 5. 生成交叉表
# =====================================
region_mechanism_ct = cross_table_one_multi(df, "region", "mechanism_type")
region_outcome_ct = cross_table_one_multi(df, "region", "outcome_type")
project_mechanism_ct = cross_table_two_multi(df, "project_type", "mechanism_type")
case_evidence_ct = pd.crosstab(df["case_level"], df["evidence_strength"])
case_evidence_ct["合计"] = case_evidence_ct.sum(axis=1)
case_evidence_ct.loc["合计"] = case_evidence_ct.sum(axis=0)

# =====================================
# 6. 生成重点案例筛选表
# =====================================
key_case_table = make_key_case_table(df)

# =====================================
# 7. 导出 Excel
# =====================================
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    region_freq.to_excel(writer, sheet_name="01_地区分布", index=False)
    case_level_freq.to_excel(writer, sheet_name="02_案例层级分布", index=False)
    evidence_freq.to_excel(writer, sheet_name="03_证据强度分布", index=False)
    project_type_freq.to_excel(writer, sheet_name="04_项目类型频数", index=False)
    beneficiaries_freq.to_excel(writer, sheet_name="05_受益对象频数", index=False)
    outcome_type_freq.to_excel(writer, sheet_name="06_成效类型频数", index=False)
    mechanism_type_freq.to_excel(writer, sheet_name="07_机制类型频数", index=False)
    small_beautiful_freq.to_excel(writer, sheet_name="08_小而美特征频数", index=False)
    region_mechanism_ct.to_excel(writer, sheet_name="09_地区x机制")
    region_outcome_ct.to_excel(writer, sheet_name="10_地区x成效")
    project_mechanism_ct.to_excel(writer, sheet_name="11_项目类型x机制")
    case_evidence_ct.to_excel(writer, sheet_name="12_层级x证据强度")
    key_case_table.to_excel(writer, sheet_name="13_重点案例筛选", index=False)

print("处理完成！结果文件生成在：", output_file)
